import io
import re
import zipfile
import html
import tempfile
import os
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

APP_TITLE = "US Foods Daily Report Generator v4"

try:
    import extract_msg
except Exception:
    extract_msg = None


def clean_text(value):
    if value is None:
        return ""
    value = html.unescape(str(value)).replace("\x00", "")
    return re.sub(r"\s+", " ", value).strip()


def fmt_date(value):
    if not value or str(value).lower() in ["nan", "nat", "none"]:
        return ""
    try:
        return pd.to_datetime(value).strftime("%m/%d/%y")
    except Exception:
        m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", str(value))
        if m:
            return f"{int(m.group(2)):02d}/{int(m.group(3)):02d}/{str(m.group(1))[2:]}"
        return clean_text(value)


def parse_date_obj(value):
    try:
        if value and str(value).lower() not in ["nan", "nat", "none"]:
            return pd.to_datetime(value).date()
    except Exception:
        pass
    return None


STOP_LABELS = [
    "JOB", "ORDER DESC", "PAGE SIZE", "COLOR", "COATING", "PAGE SETUP",
    "PAPER TYPE", "BINDERY 1", "BINDERY 2", "LAMINATE", "COLLATE",
    "QUANTITY", "RUSH", "Finishing", "Finished Size", "JobPressInstruction",
    "CUSTOMER", "SHIP", "DATE", "DESCRIPTION", "STOCK"
]


def extract_between(text, label, stop_labels):
    text = clean_text(text)
    if not text:
        return ""
    pattern = r"\bLAMINATE\b\s*:?" if label.upper() == "LAMINATE" else re.escape(label) + r"\s*:"
    m = re.search(pattern, text, flags=re.I)
    if not m:
        return ""
    start = m.end()
    stops = []
    for stop_label in stop_labels:
        sp = r"\bLAMINATE\b\s*:?" if stop_label.upper() == "LAMINATE" else re.escape(stop_label) + r"\s*:"
        sm = re.search(sp, text[start:], flags=re.I)
        if sm:
            stops.append(start + sm.start())
    end = min(stops) if stops else min(len(text), start + 200)
    return clean_text(text[start:end])


def get_full_text(blob):
    fulls = re.findall(r'<Extrinsic\s+name=["\']FullText["\']\s*>(.*?)</Extrinsic>', blob, flags=re.I | re.S)
    useful = [clean_text(x) for x in fulls if clean_text(x)]

    comments = re.findall(r'<Extrinsic\s+name=["\']Comment_\d+["\']\s*>(.*?)</Extrinsic>', blob, flags=re.I | re.S)
    comment_text = clean_text(" ".join(comments)) if comments else ""

    body_text = clean_text(blob)
    candidates = useful + [comment_text, body_text]
    candidates = [x for x in candidates if x]
    return max(candidates, key=len) if candidates else ""


def extract_spec_value(text, names):
    text = clean_text(text)
    for name in names:
        patterns = [
            rf'<SPECIFICATION\s+NAME=["\']{re.escape(name)}["\']\s*>(.*?)</SPECIFICATION>',
            rf'<Extrinsic\s+name=["\']{re.escape(name)}["\']\s*>(.*?)</Extrinsic>',
            rf'\b{re.escape(name)}\b\s*:\s*([^:]+?)(?=\s+[A-Z][A-Z0-9 /#&().-]{{2,}}\s*:|\s+Finished Size\s*:|\s+JobPressInstruction\s*:|$)',
        ]
        for pattern in patterns:
            m = re.search(pattern, text, flags=re.I | re.S)
            if m:
                return clean_text(m.group(1))
    return ""


def normalize_page_size(value):
    value = clean_text(value)
    value = re.split(r"\b(COLOR|COATING|PAGE SETUP|PAPER TYPE|BINDERY|COLLATE|QUANTITY|RUSH|LAMINATE)\b\s*: ?", value, flags=re.I)[0].strip()
    m = re.search(r"(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)", value)
    if m:
        return f"{m.group(1)} X {m.group(2)}"
    return value


def clean_paper_type(value):
    value = clean_text(value)
    value = re.split(
        r"\b(BINDERY|COATING|COLLATE|QUANTITY|RUSH|PAGE SETUP|PAGE SIZE|COLOR|Finished Size|JobPressInstruction|LAMINATE)\b\s*: ?",
        value,
        flags=re.I,
    )[0].strip(" -")
    return value


def extract_specs(spec_text, desc_fallback="", supplier_part=""):
    src = clean_text(spec_text)
    desc = clean_text(desc_fallback)
    supplier_upper = clean_text(supplier_part).upper()

    paper = extract_spec_value(src, ["PAPER TYPE", "Paper Type", "Stock"])
    if not paper:
        paper = extract_between(src, "PAPER TYPE", STOP_LABELS)

    page = extract_spec_value(src, ["PAGE SIZE", "Page Size", "Finished Size"])
    if not page:
        page = extract_between(src, "PAGE SIZE", STOP_LABELS)

    if not paper and desc:
        paper = extract_spec_value(desc, ["PAPER TYPE", "Paper Type", "Stock"])
        if not paper:
            paper = extract_between(desc, "PAPER TYPE", STOP_LABELS)

    if not page and desc:
        page = extract_spec_value(desc, ["PAGE SIZE", "Page Size", "Finished Size"])
        if not page:
            page = extract_between(desc, "PAGE SIZE", STOP_LABELS)
        if not page:
            m = re.search(r"(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)", desc)
            if m:
                page = f"{m.group(1)} X {m.group(2)}"

    paper = clean_paper_type(paper)
    page = normalize_page_size(page)

    if supplier_upper.startswith("TTSCP") and not paper:
        paper = "12pt C1S"
    if supplier_upper.startswith("TTSCP") and not page:
        page = "17 X 5"

    coating = "UV" if "UV" in paper.upper() else "NONE"

    laminate_text = extract_spec_value(src, ["LAMINATE", "Laminate", "Lamination"])
    if not laminate_text:
        laminate_text = extract_between(src, "LAMINATE", STOP_LABELS)

    lam_upper = f"{laminate_text} {src}".upper()
    laminate = "LAMINATE" if (
        re.search(r"\bLAMINATION\b|\bMATTE LAMINATE\b|\bGLOSS LAMINATE\b|\bLAMINATE\s+(MATTE|GLOSS)", lam_upper)
        or "MUST BE LAMINATED" in lam_upper
    ) else "NONE"

    return paper, page, laminate, coating


def read_msg_body_with_extract_msg(file_name, data):
    if extract_msg is None or not file_name.lower().endswith(".msg"):
        return ""

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".msg") as tmp:
            tmp.write(data)
            temp_path = tmp.name

        msg = extract_msg.Message(temp_path)
        parts = []

        for attr in ["body", "htmlBody", "subject", "date"]:
            try:
                value = getattr(msg, attr, "")
                if value:
                    parts.append(str(value))
            except Exception:
                pass

        try:
            msg.close()
        except Exception:
            pass

        return "\n".join(parts)
    except Exception:
        return ""
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass


def bytes_to_text_fast(file_name, data):
    parts = []

    msg_text = read_msg_body_with_extract_msg(file_name, data)
    if msg_text:
        parts.append(msg_text)

    for enc in ["utf-16le", "utf-8", "latin-1"]:
        try:
            decoded = data.decode(enc, errors="ignore")
            if decoded:
                parts.append(decoded)
        except Exception:
            pass

    return "\n".join(parts)


def extract_cxml_records_from_text(text, source_name=""):
    chunks = [m.group(0) for m in re.finditer(r"<cXML\b.*?</cXML>", text, flags=re.I | re.S)]

    if not chunks:
        for m in re.finditer(r'orderID=["\']([^"\']+)["\']', text, flags=re.I):
            chunks.append(text[max(0, m.start() - 5000):m.end() + 30000])

    records = []

    for chunk in chunks:
        order = re.search(r'orderID=["\']([^"\']+)["\']', chunk, flags=re.I)
        if not order:
            continue

        po = clean_text(order.group(1))

        cust_match = re.search(r'<Extrinsic\s+name=["\']custPONumber["\']\s*>(.*?)</Extrinsic>', chunk, flags=re.I | re.S)
        cust_po = clean_text(cust_match.group(1)) if cust_match else ""

        if not cust_po:
            fallback = re.search(r"\bMS\d+\b", chunk)
            cust_po = fallback.group(0) if fallback else ""

        ship_match = re.search(r'<Extrinsic\s+name=["\']dateRequired["\']\s*>(.*?)</Extrinsic>', chunk, flags=re.I | re.S)
        ship_date = clean_text(ship_match.group(1)) if ship_match else ""

        if not ship_date:
            requested = re.search(r'requestedDeliveryDate=["\']([^"\']+)["\']', chunk, flags=re.I)
            ship_date = clean_text(requested.group(1)) if requested else ""

        timestamp = re.search(r'timestamp=["\'](\d{4}-\d{2}-\d{2})', chunk, flags=re.I)
        received_date = timestamp.group(1) if timestamp else ""

        supplier = re.search(r"<SupplierPartID>(.*?)</SupplierPartID>", chunk, flags=re.I | re.S)
        supplier_part = clean_text(supplier.group(1)) if supplier else ""

        full_text = get_full_text(chunk)

        records.append({
            "PO#": po,
            "custPONumber": cust_po,
            "Ship Date": ship_date,
            "Received Date": received_date,
            "FullText": full_text,
            "SupplierPartID": supplier_part,
            "source": source_name,
            "raw": chunk,
        })

    return records


def collect_records(zip_uploads, email_uploads, progress=None):
    records = []
    files_seen = 0

    for upload in zip_uploads or []:
        if progress:
            progress.write(f"Reading ZIP: {upload.name}")

        data = upload.read()

        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = [
                n for n in zf.namelist()
                if not n.endswith("/") and n.lower().endswith((".msg", ".eml", ".xml", ".txt"))
            ]

            for i, name in enumerate(names, start=1):
                if progress:
                    progress.write(f"Processing email/XML {i} of {len(names)}: {name}")

                text = bytes_to_text_fast(name, zf.read(name))
                records.extend(extract_cxml_records_from_text(text, source_name=name))
                files_seen += 1

    for upload in email_uploads or []:
        if progress:
            progress.write(f"Processing uploaded file: {upload.name}")

        text = bytes_to_text_fast(upload.name, upload.read())
        records.extend(extract_cxml_records_from_text(text, source_name=upload.name))
        files_seen += 1

    best = {}

    for rec in records:
        po = rec["PO#"]
        score = (
            1 if rec.get("FullText") else 0,
            1 if rec.get("custPONumber") else 0,
            1 if rec.get("Ship Date") else 0,
            1 if rec.get("SupplierPartID") else 0,
            len(rec.get("raw", "")),
        )

        if po not in best or score > best[po][0]:
            best[po] = (score, rec)

    return {po: rec for po, (score, rec) in best.items()}, files_seen, len(records)


def load_tracking(upload):
    data = upload.read()

    try:
        tables = pd.read_html(io.BytesIO(data))
        df = tables[0]
        df.columns = list(df.iloc[0])
        df = df.iloc[1:].copy()
    except Exception:
        df = pd.read_excel(io.BytesIO(data))

    df.columns = [clean_text(c) for c in df.columns]

    for col in df.columns:
        df[col] = df[col].apply(lambda x: "" if pd.isna(x) else x)

    needed = ["Job No", "Order Description", "PO#", "Tracking Number"]
    missing = [x for x in needed if x not in df.columns]

    if missing:
        raise ValueError(f"Tracking report missing columns: {missing}")

    if "Desc" not in df.columns:
        df["Desc"] = ""

    return df[["Job No", "Order Description", "PO#", "Tracking Number", "Desc"]].copy()


def load_cancelled(upload):
    data = upload.read()
    df = pd.read_excel(io.BytesIO(data))
    df.columns = [clean_text(c) for c in df.columns]

    job_col = "Job #" if "Job #" in df.columns else next((c for c in df.columns if "job" in c.lower()), None)
    status_col = "Status" if "Status" in df.columns else next((c for c in df.columns if "status" in c.lower()), None)

    if not job_col or not status_col:
        return set()

    cancelled_df = df[df[status_col].astype(str).str.contains("cancel", case=False, na=False)]
    return set(cancelled_df[job_col].astype(str).str.strip())


def make_report_row(row, rec, email_found):
    spec_source = f"{rec.get('FullText', '')} {rec.get('raw', '')}" if rec else ""

    paper, page, laminate, coating = extract_specs(
        spec_source,
        clean_text(row.get("Desc", "")),
        rec.get("SupplierPartID", "") if rec else "",
    )

    if not email_found:
        paper = ""
        page = ""
        laminate = "NONE"
        coating = "NONE"

    return {
        "Job No": clean_text(row["Job No"]),
        "Order Description": clean_text(row["Order Description"]),
        "PO#": clean_text(row["PO#"]),
        "custPONumber": clean_text(rec.get("custPONumber", "")) if rec else "",
        "Email Found": "YES" if email_found else "NO",
        "Ship Date": fmt_date(rec.get("Ship Date", "")) if rec else "",
        "Received Date": fmt_date(rec.get("Received Date", "")) if rec else "",
        "Paper Type": paper,
        "Page Size": page,
        "Laminate": laminate,
        "Coating": coating,
    }


def sort_report(df):
    if df.empty:
        return df

    def priority(r):
        if r["Laminate"] == "LAMINATE":
            return 0
        if r["Coating"] == "UV":
            return 1
        return 2

    df = df.copy()
    df["_priority"] = df.apply(priority, axis=1)
    df["_ship_sort"] = pd.to_datetime(df["Ship Date"], errors="coerce")
    df = df.sort_values(["_priority", "_ship_sort", "PO#"], na_position="last")

    return df.drop(columns=["_priority", "_ship_sort"])


def build_report(tracking_upload, cancel_upload, zip_uploads, email_uploads, progress):
    progress.write("Loading tracking report...")
    master = load_tracking(tracking_upload)
    original_rows = len(master)

    progress.write("Removing rows with tracking numbers...")
    master = master[master["Tracking Number"].astype(str).str.strip().eq("")].copy()
    untracked_rows = len(master)

    progress.write("Loading cancelled status report...")
    cancelled = load_cancelled(cancel_upload)
    master = master[~master["Job No"].astype(str).str.strip().isin(cancelled)].copy()
    after_cancel = len(master)

    progress.write("Reading Outlook email bodies/XML...")
    po_map, files_seen, xml_records = collect_records(zip_uploads, email_uploads, progress)

    rows = []
    missing_email_rows = []
    future_rows = []
    today = date.today()

    progress.write("Merging XML data into open jobs...")

    for _, row in master.iterrows():
        po = clean_text(row["PO#"])
        rec = po_map.get(po)
        email_found = rec is not None

        report_row = make_report_row(row, rec or {}, email_found)

        ship_obj = parse_date_obj(rec.get("Ship Date", "") if rec else "")

        if ship_obj and ship_obj > today:
            future_row = report_row.copy()
            future_row["Reason Excluded"] = "Ship Date is after today"
            future_rows.append(future_row)
            continue

        rows.append(report_row)

        if not email_found:
            missing_email_rows.append({
                "Job No": report_row["Job No"],
                "Order Description": report_row["Order Description"],
                "PO#": report_row["PO#"],
                "Email Found": "NO",
            })

    columns = [
        "Job No", "Order Description", "PO#", "custPONumber", "Email Found",
        "Ship Date", "Received Date", "Paper Type", "Page Size", "Laminate", "Coating"
    ]

    final = pd.DataFrame(rows, columns=columns)
    final = sort_report(final)

    missing_emails = pd.DataFrame(
        missing_email_rows,
        columns=["Job No", "Order Description", "PO#", "Email Found"],
    )

    future_exclusions = pd.DataFrame(
        future_rows,
        columns=columns + ["Reason Excluded"],
    )

    stats = {
        "Original tracking rows": original_rows,
        "Rows after tracking removal": untracked_rows,
        "Rows after cancelled removal": after_cancel,
        "Email/XML files scanned": files_seen,
        "Embedded XML orders found": len(po_map),
        "Raw XML records found": xml_records,
        "Final report rows": len(final),
        "Missing Email Count": len(missing_emails),
        "Future Ship Date Exclusions": len(future_exclusions),
        "Open jobs accounted for": len(final) + len(future_exclusions),
    }

    expected_jobs = set(master["Job No"].astype(str).str.strip())
    final_jobs = set(final["Job No"].astype(str).str.strip()) if not final.empty else set()
    future_jobs = set(future_exclusions["Job No"].astype(str).str.strip()) if not future_exclusions.empty else set()

    missing_from_output = sorted(expected_jobs - final_jobs - future_jobs)
    stats["Validation Missing Job Count"] = len(missing_from_output)

    if missing_from_output:
        raise ValueError(
            "Validation failed. These open, non-cancelled jobs were not accounted for: "
            + ", ".join(missing_from_output[:50])
            + (" ..." if len(missing_from_output) > 50 else "")
        )

    return final, missing_emails, future_exclusions, stats


def format_worksheet(ws):
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    duplicate_fill = PatternFill("solid", fgColor="FFF2CC")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    cust_col = None

    for cell in ws[1]:
        if clean_text(cell.value) == "custPONumber":
            cust_col = cell.column
            break

    if cust_col:
        values = {}

        for row_num in range(2, ws.max_row + 1):
            value = ws.cell(row_num, cust_col).value
            if value:
                values.setdefault(value, []).append(row_num)

        for row_nums in values.values():
            if len(row_nums) > 1:
                for row_num in row_nums:
                    ws.cell(row_num, cust_col).fill = duplicate_fill

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)

            if clean_text(ws.cell(1, cell.column).value) in ["Ship Date", "Received Date"]:
                cell.number_format = "MM/DD/YY"

    ws.freeze_panes = "A2"

    for col_idx in range(1, ws.max_column + 1):
        max_len = max(len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row + 1))
        width = min(max(max_len + 2, 10), 60)
        header = clean_text(ws.cell(1, col_idx).value)

        if header in ["Order Description", "Paper Type"]:
            width = min(max(width, 30), 60)

        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_excel(final_df, missing_emails_df, future_exclusions_df, stats):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name="Filled Report", index=False)
        final_df[final_df["Laminate"].eq("LAMINATE")].to_excel(writer, sheet_name="LAMINATE", index=False)
        final_df[final_df["Coating"].eq("UV")].to_excel(writer, sheet_name="UV COATING", index=False)
        final_df[(final_df["Laminate"].eq("NONE")) & (final_df["Coating"].eq("NONE"))].to_excel(writer, sheet_name="TRIM TO SIZE", index=False)
        missing_emails_df.to_excel(writer, sheet_name="MISSING EMAILS", index=False)
        future_exclusions_df.to_excel(writer, sheet_name="FUTURE SHIP DATE EXCLUSIONS", index=False)

        summary_df = pd.DataFrame(list(stats.items()), columns=["Metric", "Value"])
        summary_df.to_excel(writer, sheet_name="SUMMARY", index=False)

    output.seek(0)
    wb = load_workbook(output)

    for ws in wb.worksheets:
        format_worksheet(ws)

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output.getvalue()


st.set_page_config(page_title=APP_TITLE, page_icon="📊", layout="wide")
st.title(APP_TITLE)

st.info(
    "v4: Master report is the source of truth. Open/non-cancelled jobs stay in the report even when no email/XML is found. "
    "Adds Email Found, MISSING EMAILS, FUTURE SHIP DATE EXCLUSIONS, and validation."
)

left, right = st.columns(2)

with left:
    tracking_upload = st.file_uploader("1. Master Tracking Numbers Report", type=["xls", "xlsx"])
    cancel_upload = st.file_uploader("2. Cancelled Status Report", type=["xlsx", "xls"])

with right:
    zip_uploads = st.file_uploader("3. US Foods Outlook ZIP file(s)", type=["zip"], accept_multiple_files=True)
    email_uploads = st.file_uploader(
        "Optional: individual .msg, .eml, .xml, .txt files",
        type=["msg", "eml", "xml", "txt"],
        accept_multiple_files=True,
    )

if st.button("Generate Report", type="primary"):
    status = st.empty()

    with status.container():
        st.write("Starting report...")

    if not tracking_upload:
        st.error("Upload the Master Tracking Numbers Report.")
    elif not cancel_upload:
        st.error("Upload the Cancelled Status Report.")
    elif not zip_uploads and not email_uploads:
        st.error("Upload either a US Foods ZIP or individual email/XML files.")
    else:
        try:
            final_df, missing_emails_df, future_exclusions_df, stats = build_report(
                tracking_upload, cancel_upload, zip_uploads, email_uploads, status
            )

            report_bytes = write_excel(final_df, missing_emails_df, future_exclusions_df, stats)

            st.success("Report generated successfully.")

            cols = st.columns(5)
            cols[0].metric("Final Rows", stats["Final report rows"])
            cols[1].metric("Missing Emails", stats["Missing Email Count"])
            cols[2].metric("Future Exclusions", stats["Future Ship Date Exclusions"])
            cols[3].metric("XML Orders Found", stats["Embedded XML orders found"])
            cols[4].metric("Files Scanned", stats["Email/XML files scanned"])

            if stats["Missing Email Count"]:
                st.warning(
                    f"{stats['Missing Email Count']} open jobs were retained in the report but no matching Outlook email/XML data was found."
                )

            if stats["Future Ship Date Exclusions"]:
                st.info(
                    f"{stats['Future Ship Date Exclusions']} open jobs were excluded because their Ship Date is after today."
                )

            with st.expander("Processing Stats"):
                st.json(stats)

            blank_counts = final_df.eq("").sum()

            if blank_counts.sum() > 0:
                st.warning("Some blanks were detected. Check Email Found and MISSING EMAILS for missing XML/email matches.")
                st.dataframe(blank_counts.rename("Blank Count"))
            else:
                st.info("No blank fields detected.")

            if not missing_emails_df.empty:
                with st.expander("Missing Emails"):
                    st.dataframe(missing_emails_df, use_container_width=True)

            if not future_exclusions_df.empty:
                with st.expander("Future Ship Date Exclusions"):
                    st.dataframe(future_exclusions_df, use_container_width=True)

            st.download_button(
                "Download Excel Report",
                data=report_bytes,
                file_name=f"US_Foods_Report_{date.today().strftime('%m%d%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            st.subheader("Preview")
            st.dataframe(final_df, use_container_width=True)

        except Exception as e:
            st.error(f"Report failed: {e}")
            st.write("Screenshot this error or copy it to Peter.")